#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 21 13:30:49 2020

@author: EthanMorse
"""

# for .csv, .xlsx reading/converting
from openpyxl import load_workbook





def create_html_table():

    wb = load_workbook(filename = "bucket_list.xlsx")
    ws = wb.active
    num_row = ws.max_row

    start_table = "<table>"
    end_table = "</table>"
    final_table = ""
    header = "<tr><th>Item</th><th>Category</th><th>Country</th><th>City</th><th>Description</th><th>Dates</th><th>Priority (1 (high) - 5)</th><th>Notes</th><th>Complete?</th></tr>"
    all_rows = ""


    for row in range(2, num_row+1):

        # find specific cell values for information
        item_cell = "A" + str(row)
        category_cell = "B" + str(row)
        country_cell = "C" + str(row)
        city_cell = "D" + str(row)
        description_cell = "E" + str(row) 
        dates_cell = "F" + str(row)
        priority_cell = "G" + str(row)
        notes_cell = "H" + str(row)
        complete_cell = "I" + str(row)


        # create HTML table strings containing each cell's info
        start_row = """<tr align="middle">"""
        item = "<td>" + str(ws[item_cell].value) + "</td>"
        category = "<td>" + str(ws[category_cell].value) + "</td>"
        country = "<td>" + str(ws[country_cell].value) + "</td>"
        city = "<td>" + str(ws[city_cell].value) + "</td>"
        description = "<td>" + str(ws[description_cell].value) + "</td>"
        dates = "<td>" + str(ws[dates_cell].value) + "</td>"
        priority = "<td>" + str(ws[priority_cell].value) + "</td>"
        notes = "<td>" + str(ws[notes_cell].value) + "</td>"
        complete = "<td>" + str(ws[complete_cell].value) + "</td>"
        end_row = "</tr>"

        # add all info to create a single HTML table row
        all_rows = all_rows + start_row + item + category + country + city + description + dates + priority + notes + complete + end_row
    
    # concatenate all strings to create final table
    final_table = start_table + header + all_rows + end_table

    return final_table



def delete_append_to_file(final_table):

    # open html file for reading, assign contents to variable lines
    with open("bucket_list.html", "r") as html_file:
        lines = html_file.readlines()

    # open file for writing, check if "<table>" is in line.
    # If yes, effectively delete by writing all other lines
    with open("bucket_list.html", "w") as html_file:
        for line in lines:
            if "<table>" not in line:
                html_file.write(line)

    # close file before appending
    html_file.close()

    # open file for appending
    html_file = open("bucket_list.html", "a")

    # create string with all html info, include closing html tags
    file_contents = final_table + "</body></html>"
    print(file_contents)

    # append newest table to end of file
    html_file.write(file_contents)



def main():
    
    final_table = create_html_table()
    
    delete_append_to_file(final_table)
    
    

main()
    