#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on 5 April 2020
Updated on 14 April 2020

@author: EthanMorse

- download goodreads spreadsheet
- convert to .xlsx
- go through read/not-read column. if there is data, add row number to list
- add title, author, rating, and review to html table
    - https://www.csestack.org/python-generate-html/
- write to book_reviews.html file (ensure all the other info has been added
beforehand)

To update ChromeDriver:
- bring up Alfred search, "find ChromeDriver": /Users/EthanMorse/opt/anaconda3/lib/python3.7/site-packages
- download correct ChromeDriver (CD) packages: https://chromedriver.chromium.org/downloads
- unzip
- replace old CD executable with new CD executable
- give CD permissions in System Preferences

"""

# selenium packages
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

# file path to goodreads info
import sys
sys.path.append("/Users/EthanMorse/Documents/personal/Python")
from goodreads_info import em, pw

# for .csv, .xlsx reading/converting
from openpyxl import load_workbook
import pandas as pd

# sleeping, deleting files
import time
import os
import os.path
from os import path



# =============================================================================
# export_library function:
#   - sign in to goodreads
#   - export, download library
# =============================================================================

def export_library():

    # open Chrome webdriver
    driver = webdriver.Chrome(executable_path="/Users/EthanMorse/opt/anaconda3/lib/python3.7/site-packages/chromedriver")

    # go to Goodreads sign in page
    driver.get("https://www.goodreads.com/user/sign_in")

    # find buttons for email, pw, login
    user_email = driver.find_element_by_id("user_email")
    user_pw = driver.find_element_by_id("user_password")
    login_button = driver.find_element_by_name("next")

    # enter login info, log in
    user_email.send_keys(em)
    user_pw.send_keys(pw)
    login_button.send_keys(Keys.RETURN)

    # navigate to webpage, wait until fully open
    driver.get("https://www.goodreads.com/review/import")

    # find export button
    export_button = driver.find_element_by_class_name("gr-form--compact__submitButton")

    # press enter on export button
    export_button.send_keys(Keys.RETURN)

    # wait for library to be exported
    time.sleep(120)

    # go to link
    driver.get("https://www.goodreads.com/review_porter/export/75258150/goodreads_export.csv")

    # wait 10 sec before closing, just in case
    time.sleep(5)

    # close Chrome webdriver
    driver.close()


# =============================================================================
# csv_to_xlsx function:
#   - delete existing goodreads library export
#   - convert csv to xlsx
#   - delete newly-downloaded csv
# =============================================================================

def csv_to_xlsx():

    # delete existing goodreads export to prevent overwriting
    if path.exists("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/media/goodreads_library_export.xlsx"):
        os.remove("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/media/goodreads_library_export.xlsx")

    # rewrite .csv file contents to .xlsx file using pandas and xlsxwriter engine
    csv_file = pd.read_csv("/Users/EthanMorse/Downloads/goodreads_library_export.csv")
    csv_file.to_excel("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/media/goodreads_library_export.xlsx", engine = "xlsxwriter")

    # delete .csv file to prevent future overwriting
    os.remove("/Users/EthanMorse/Downloads/goodreads_library_export.csv")

# =============================================================================
# get_rows function:
#   - go through column S, which indicates if I've read the book or not
#   - if I've read it, store the row number for create_html_table
# =============================================================================

def get_rows():

    # load workbook
    wb = load_workbook(filename = "goodreads_library_export.xlsx")
    ws = wb.active

    # initialize row number to be 1, no applicable rows in list yet
    row_num = 1
    row_list = []

    # iterate through "read" or "to-read" column (column letter S)
    for cell in ws["T"]:

        #if already red, append row number to row_list
        if cell.value == ("read"):
            row_list.append(row_num)
        row_num += 1

    return row_list


# =============================================================================
# create_html_table function:
#   - find title/author/rating/review cells
#   - format cells into html table syntax, add to master string (all_rows)
#   - create final_table by adding beginning/end of table and header row
# =============================================================================

def create_html_table(row_list):

    wb = load_workbook(filename = "goodreads_library_export.xlsx")
    ws = wb.active

    start_table = "<table>"
    end_table = "</table>"
    final_table = ""
    header = "<tr><th>Title</th><th>Author</th><th>My Rating</th><th>My Review</th></tr>"
    all_rows = ""


    for row_num in row_list:

        # find specific cell values for title, author, rating, review
        title_cell = "C" + str(row_num)
        author_cell = "D" + str(row_num)
        my_rating_cell = "I" + str(row_num)
        my_review_cell = "U" + str(row_num)

        # create HTML table strings containing each cell's info
        start_row = "<tr>"
        title = "<td>" + str(ws[title_cell].value) + "</td>"
        author = "<td>" + str(ws[author_cell].value) + "</td>"
        my_rating = "<td>" + str(ws[my_rating_cell].value) + "</td>"
        my_review = "<td>" + str(ws[my_review_cell].value) + "</td>"
        end_row = "</tr>"

        # add all info to create a single HTML table row
        all_rows = all_rows + start_row + title + author + my_rating + my_review + end_row

    # concatenate all strings to create final table
    final_table = start_table + header + all_rows + end_table

    return final_table


# =============================================================================
# delete_append_to_file function:
#   - delete old table
#   - append new table
# =============================================================================

def delete_append_to_file(final_table):

    # open html file for reading, assign contents to variable lines
    with open("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/books.html", "r") as html_file:
        lines = html_file.readlines()

    # open file for writing, check if "<table>" is in line.
    # If yes, effectively delete by writing all other lines
    with open("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/books.html", "w") as html_file:
        for line in lines:
            if "<table>" not in line:
                html_file.write(line)

    # close file before appending
    html_file.close()

    # open file for appending
    html_file = open("/Users/EthanMorse/Documents/personal/website/ethanmorse.github.io/knowledge/books.html", "a")

    # create string with all html info, include closing html tags
    file_contents = final_table + "</body></html>"

    # append newest table to end of file
    html_file.write(file_contents)


# =============================================================================
# main function:
#   - export current library
#   - convert .csv to .xlsx
#   - get list of row numbers of read books
#   - create html table with all read books
#   - update html table in book_reviews.html
# =============================================================================

def main():

    #export_library()

    csv_to_xlsx()

    row_list = get_rows()

    final_table = create_html_table(row_list)

    delete_append_to_file(final_table)

# run main() program
main()
