#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Apr  5 19:12:07 2020

@author: EthanMorse

- open Goodreads book spreadsheet
- go through reviews column. if there is data, add row number to list
- add title, author, rating, and review to html table
    - https://www.csestack.org/python-generate-html/
- write to book_reviews.html file (ensure all the other info has been added
beforehand)

"""

from openpyxl import load_workbook

# =============================================================================
# get_rows function:
#   - go through column S, which indicates if I've read the book or not
#   - if I've read it, store the row number for create_html_table
# =============================================================================

def get_rows():    
    
    # load workbook
    wb = load_workbook(filename = "goodreads_library_export.xlsx")
    ws = wb.active
    
    # initialize row number to be 1, no applicable rows in list yet
    row_num = 1
    row_list = []
    
    # iterate through "read" or "to-read" column (column letter S)
    for cell in ws["S"]:
        
        #if I've read it, append row number to row_list
        if cell.value == ("read"):
            row_list.append(row_num)
        row_num += 1
        
    return row_list

        
# =============================================================================
# create_html_table function:
#   - find title/author/rating/review cells
#   - format cells into html table syntax, add to master string (all_rows) 
#   - create final_table by adding beginning/end of table and header row
# =============================================================================

def create_html_table(row_list):
    
    wb = load_workbook(filename = "goodreads_library_export.xlsx")
    ws = wb.active
    
    start_table = "<table>"
    end_table = "</table>"
    final_table = ""
    header = "<tr><th>Title</th><th>Author</th><th>My Rating</th><th>My Review</th></tr>"
    all_rows = ""
    
    
    for row_num in row_list:
        
        # find specific cell values for title, author, rating, review
        title_cell = "B" + str(row_num)
        author_cell = "C" + str(row_num)
        my_rating_cell = "H" + str(row_num)
        my_review_cell = "T" + str(row_num)
        
        # create HTML table strings containing each cell's info
        start_row = "<tr>"
        title = "<td>" + str(ws[title_cell].value) + "</td>"
        author = "<td>" + str(ws[author_cell].value) + "</td>"
        my_rating = "<td>" + str(ws[my_rating_cell].value) + "</td>"
        my_review = "<td>" + str(ws[my_review_cell].value) + "</td>"
        end_row = "</tr>"
        
        # add all info to create a single HTML table row
        all_rows = all_rows + start_row + title + author + my_rating + my_review + end_row
    
    # concatenate all strings to create final table
    final_table = start_table + header + all_rows + end_table

    return final_table


# =============================================================================
# delete_append_to_file function:
#   - delete old table
#   - append new table
# =============================================================================

def delete_append_to_file(final_table):
    
    # open html file for reading, assign contents to variable lines
    with open("book_reviews.html", "r") as html_file:
        lines = html_file.readlines()
        
    # open file for writing, check if "<table>" is in line.
    # If yes, effectively delete by writing all other lines
    with open("book_reviews.html", "w") as html_file:
        for line in lines:
            if "<table>" not in line:
                html_file.write(line)
    
    # close file before appending     
    html_file.close()
    
    # open file for appending
    html_file = open("book_reviews.html", "a")
    
    # create string with all html info, include closing html tags
    file_contents = final_table + "</body></html>"
   
    # append newest table to end of file
    html_file.write(file_contents)


# =============================================================================
# main function:
#   - get list of row numbers of read books
#   - create html table with all read books
#   - update html table in book_reviews.html
# =============================================================================
    
def main():
    
    row_list = get_rows()
    
    final_table = create_html_table(row_list)

    delete_append_to_file(final_table)
    
# run main() program
main()
    
    
    
